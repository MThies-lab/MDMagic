#!/usr/bin/env python3
"""
Document converter for MarkdownMagic - simplified version for testing
Converts various document formats to Markdown
"""

import os
import re
from pathlib import Path

class DocumentConverter:
    """
    Converts various document formats to Markdown with structure preservation
    """
    
    def __init__(self, tesseract_path=None, enable_ai=True):
        """Initialize the document converter with AI features"""
        self.supported_formats = {
            '.txt': self._txt_to_md,
            '.pdf': self._pdf_to_md,
            '.docx': self._docx_to_md,
            '.odt': self._odt_to_md,
            '.rtf': self._rtf_to_md,
            '.html': self._html_to_md,
            '.htm': self._html_to_md,
            '.xlsx': self._xlsx_to_md,
            '.xls': self._xlsx_to_md,
            '.png': self._image_to_md,
            '.jpg': self._image_to_md,
            '.jpeg': self._image_to_md,
            '.gif': self._image_to_md,
            '.bmp': self._image_to_md,
            '.tiff': self._image_to_md,
            '.tif': self._image_to_md
        }
        self.tesseract_path = tesseract_path
        self.enable_ai = enable_ai
        
        # Initialize AI vision processor directly for better integration
        try:
            from ai_vision_processor import AIVisionProcessor
            self.image_processor = AIVisionProcessor(tesseract_path=tesseract_path, enable_ai=enable_ai)
            if enable_ai:
                print("MarkdownMagic Document Converter initialized with AI-powered image processing")
            else:
                print("MarkdownMagic Document Converter initialized with basic image processing")
        except ImportError:
            # Fallback to basic ImageProcessor if AI not available
            try:
                from image_processor import ImageProcessor
                self.image_processor = ImageProcessor(tesseract_path=tesseract_path, enable_ai=False)
                print("MarkdownMagic Document Converter initialized with basic image processing")
            except ImportError:
                self.image_processor = None
                print("MarkdownMagic Document Converter initialized (image processing unavailable)")
    
    def convert_to_markdown(self, input_file, output_file=None):
        """
        Convert a document to Markdown format
        
        Args:
            input_file (str): Path to input document
            output_file (str, optional): Path for output Markdown file
            
        Returns:
            str: Path to created Markdown file
        """
        
        # Validate input file
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Get file extension
        file_ext = Path(input_file).suffix.lower()
        
        # Check if format is supported
        if file_ext not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        # Generate output filename if not provided
        if output_file is None:
            input_path = Path(input_file)
            output_file = str(input_path.with_suffix('.md'))
        
        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Convert using appropriate method
        converter_method = self.supported_formats[file_ext]
        converter_method(input_file, output_file)
        
        return output_file
    
    def _txt_to_md(self, input_file, output_file):
        """Convert plain text to Markdown with intelligent structure detection"""
        print(f"Converting TXT to Markdown: {input_file}")
        
        with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Convert plain text to structured Markdown
            markdown_content = self._enhance_plain_text_to_markdown(content)
            f.write(markdown_content)
        
        print(f"✓ TXT to Markdown conversion complete: {output_file}")
        return output_file

    def _pdf_to_md(self, input_file, output_file):
        """Convert PDF to Markdown with image extraction"""
        print(f"Converting PDF to Markdown: {input_file}")
        
        try:
            import fitz  # PyMuPDF
            pymupdf_available = True
        except ImportError:
            pymupdf_available = False
            print("PyMuPDF not installed. Install with: pip3 install PyMuPDF")
        
        # If we have PyMuPDF and image processor, use enhanced conversion
        if pymupdf_available and self.image_processor:
            try:
                # Open the PDF
                pdf = fitz.open(input_file)
                
                # Create image folder for this document
                images_folder = self.image_processor.create_image_folder(output_file)
                
                # Track document structure
                image_count = 0
                total_pages = len(pdf)
                
                # Track potential headers/footers to remove
                page_texts = []
                potential_headers = []
                potential_footers = []
                
                # First pass: collect all page texts to identify headers/footers
                for page_num in range(total_pages):
                    page = pdf[page_num]
                    text = page.get_text()
                    page_texts.append(text)
                    
                    # Split into lines for header/footer detection
                    lines = text.strip().split('\n')
                    if len(lines) > 2:
                        # Potential header (first few lines)
                        potential_headers.append(lines[0:3])
                        # Potential footer (last few lines)  
                        potential_footers.append(lines[-3:])
                
                # Identify common headers/footers that appear on multiple pages
                common_headers = self._find_common_lines(potential_headers) if len(potential_headers) > 1 else []
                common_footers = self._find_common_lines(potential_footers) if len(potential_footers) > 1 else []
                
                # Prepare to extract text with structure
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Add title
                    f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # Process each page
                    for page_num in range(total_pages):
                        print(f"Processing page {page_num + 1}...")
                        
                        # Get the page
                        page = pdf[page_num]
                        
                        # Use pre-collected text and clean it
                        text = page_texts[page_num]
                        cleaned_text = self._remove_headers_footers(text, common_headers, common_footers)
                        
                        # Extract images from this page
                        try:
                            image_list = page.get_images()
                            for img_index, img in enumerate(image_list):
                                try:
                                    # Extract image
                                    xref = img[0]
                                    base_image = pdf.extract_image(xref)
                                    image_bytes = base_image["image"]
                                    
                                    image_count += 1
                                    
                                    # Process image with OCR
                                    page_info = f"pg{page_num + 1}"
                                    markdown_placeholder = self.image_processor.process_image(
                                        image_bytes, image_count, images_folder, page_info
                                    )
                                    
                                    # Add to markdown
                                    f.write(f"{markdown_placeholder}\n\n")
                                    
                                except Exception as e:
                                    print(f"  ⚠ Failed to extract image {img_index + 1}: {e}")
                                    # Create a placeholder anyway
                                    placeholder = f"![Image {image_count + 1}, pg{page_num + 1} - Could not extract image]({images_folder}/image_{image_count + 1}.png)"
                                    f.write(f"{placeholder}\n\n")
                                    image_count += 1
                        
                        except Exception as e:
                            print(f"  ⚠ Error processing images on page {page_num + 1}: {e}")
                        
                        # Add cleaned page text with Markdown enhancement
                        if cleaned_text.strip():
                            enhanced_text = self._enhance_plain_text_to_markdown(cleaned_text.strip())
                            f.write(f"{enhanced_text}\n\n")
                        
                        # Add page break except for last page
                        if page_num < total_pages - 1:
                            f.write("\n---\n\n")
                
                # Close the PDF properly
                pdf.close()
                
                if image_count > 0:
                    print(f"✓ PDF conversion complete with {total_pages} pages and {image_count} images processed")
                else:
                    print(f"✓ PDF conversion complete with {total_pages} pages processed")
                
                return output_file
                
            except Exception as e:
                print(f"! Error during enhanced PDF conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is a PDF file converted to Markdown*\n\n")
            
            if not pymupdf_available:
                f.write("For full PDF conversion with image extraction, please ensure PyMuPDF is installed:\n")
                f.write("```\npip3 install PyMuPDF\n```\n\n")
            elif not self.image_processor:
                f.write("For image extraction, please ensure image_processor.py is available\n\n")
            
        print(f"✓ PDF to Markdown conversion complete: {output_file}")
        return output_file

    def _docx_to_md(self, input_file, output_file):
        """Convert DOCX to Markdown with image extraction"""
        print(f"Converting DOCX to Markdown: {input_file}")
        
        try:
            from docx import Document
            python_docx_available = True
        except ImportError:
            python_docx_available = False
            print("python-docx not installed. Install with: pip3 install python-docx")
        
        # If we have python-docx and image processor, use enhanced conversion
        if python_docx_available and self.image_processor:
            try:
                # Open the document
                doc = Document(input_file)
                
                # Create image folder
                images_folder = self.image_processor.create_image_folder(output_file)
                
                image_count = 0
                
                with open(output_file, 'w', encoding='utf-8') as f:
                    # YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Don't add redundant title - let document content provide its own headings
                    
                    # First pass: extract images from document relationships
                    try:
                        # Extract images from document relationships
                        if hasattr(doc, 'part') and hasattr(doc.part, 'rels'):
                            for rel_id, rel in doc.part.rels.items():
                                if "image" in rel.target_ref:
                                    try:
                                        # Get image data
                                        image_part = rel.target_part
                                        image_bytes = image_part.blob
                                        
                                        image_count += 1
                                        
                                        # Process image with OCR
                                        markdown_placeholder = self.image_processor.process_image(
                                            image_bytes, image_count, images_folder, f"pos{image_count}"
                                        )
                                        
                                        print(f"Extracted image {image_count} from DOCX")
                                        
                                    except Exception as e:
                                        print(f"Warning: Could not extract image {rel_id}: {e}")
                    except Exception as e:
                        print(f"Warning: Could not extract images from DOCX: {e}")
                    
                    # Second pass: process document content
                    image_refs_used = 0
                    
                    for paragraph in doc.paragraphs:
                        text = paragraph.text.strip()
                        
                        # Check if paragraph contains an image
                        has_inline_image = False
                        for run in paragraph.runs:
                            if hasattr(run, 'element') and run.element.findall('.//a:blip', namespaces={'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}):
                                has_inline_image = True
                                break
                        
                        # If paragraph has an image, insert the image placeholder
                        if has_inline_image and image_refs_used < image_count:
                            image_refs_used += 1
                            markdown_placeholder = self.image_processor.create_markdown_placeholder(
                                image_refs_used, f"pos{image_refs_used}", 
                                f"Image {image_refs_used} from document", images_folder
                            )
                            f.write(f"{markdown_placeholder}\n\n")
                            
                            # Also include any text in the paragraph
                            if text:
                                f.write(f"{text}\n\n")
                            continue
                        
                        if not text:
                            f.write('\n')
                            continue
                        
                        # Check paragraph style for headings
                        style_name = paragraph.style.name.lower()
                        
                        if 'heading' in style_name:
                            # Extract heading level
                            if 'heading 1' in style_name:
                                f.write(f"# {text}\n\n")
                            elif 'heading 2' in style_name:
                                f.write(f"## {text}\n\n")
                            elif 'heading 3' in style_name:
                                f.write(f"### {text}\n\n")
                            elif 'heading 4' in style_name:
                                f.write(f"#### {text}\n\n")
                            elif 'heading 5' in style_name:
                                f.write(f"##### {text}\n\n")
                            elif 'heading 6' in style_name:
                                f.write(f"###### {text}\n\n")
                            else:
                                f.write(f"## {text}\n\n")
                            continue
                        
                        # Process text formatting
                        processed_text = ""
                        for run in paragraph.runs:
                            run_text = run.text
                            if run.bold:
                                run_text = f"**{run_text}**"
                            if run.italic:
                                run_text = f"*{run_text}*"
                            processed_text += run_text
                        
                        # Regular paragraph
                        f.write(f"{processed_text}\n\n")
                    
                    # Add any remaining images that weren't inline
                    while image_refs_used < image_count:
                        image_refs_used += 1
                        markdown_placeholder = self.image_processor.create_markdown_placeholder(
                            image_refs_used, f"pos{image_refs_used}", 
                            f"Image {image_refs_used} from document", images_folder
                        )
                        f.write(f"{markdown_placeholder}\n\n")
                
                if image_count > 0:
                    print(f"✓ DOCX to Markdown conversion complete with {image_count} images processed")
                else:
                    print("✓ DOCX to Markdown conversion complete")
                
                return output_file
                
            except Exception as e:
                print(f"! Error during enhanced DOCX conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is a DOCX file converted to Markdown*\n\n")
            
            if not python_docx_available:
                f.write("For full DOCX conversion with image extraction, please ensure python-docx is installed:\n")
                f.write("```\npip3 install python-docx\n```\n\n")
            elif not self.image_processor:
                f.write("For image extraction, please ensure image_processor.py is available\n\n")
            
        print(f"✓ DOCX to Markdown conversion complete: {output_file}")
        return output_file

    def _odt_to_md(self, input_file, output_file):
        """Convert ODT to Markdown using pypandoc"""
        print(f"Converting ODT to Markdown: {input_file}")
        
        try:
            import pypandoc
            pypandoc_available = True
        except ImportError:
            pypandoc_available = False
            print("pypandoc not installed. Install with: pip3 install pypandoc")
        
        if pypandoc_available:
            try:
                # Convert ODT directly to markdown content using pypandoc
                markdown_content = pypandoc.convert_file(input_file, 'markdown')
                
                # Write to output file with frontmatter
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Don't add redundant title - pypandoc preserves document structure
                    # Write the converted content
                    f.write(markdown_content)
                
                print(f"✓ ODT to Markdown conversion complete: {output_file}")
                return output_file
                
            except Exception as e:
                print(f"! Error during ODT conversion with pypandoc: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an ODT file converted to Markdown*\n\n")
            
            if not pypandoc_available:
                f.write("For full ODT conversion, please ensure pypandoc is installed:\n")
                f.write("```\npip3 install pypandoc\n```\n\n")
                f.write("You may also need pandoc installed on your system.\n\n")
            
        print(f"✓ ODT to Markdown conversion complete: {output_file}")
        return output_file

    def _rtf_to_md(self, input_file, output_file):
        """Simple RTF to Markdown conversion for testing"""
        print(f"Converting RTF to Markdown: {input_file}")
        
        try:
            from striprtf.striprtf import rtf_to_text
            rtf_to_text_available = True
        except ImportError:
            rtf_to_text_available = False
            print("striprtf not installed. Install with: pip3 install striprtf")
        
        if rtf_to_text_available:
            try:
                # Read RTF content
                with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
                    rtf_content = f.read()
                
                # Convert RTF to plain text
                text_content = rtf_to_text(rtf_content)
                
                # Write to markdown
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Don't add redundant title - let document content provide its own structure
                    # Enhance plain text to structured Markdown
                    enhanced_content = self._enhance_plain_text_to_markdown(text_content)
                    f.write(enhanced_content)
                
                print(f"✓ RTF to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during RTF conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an RTF file converted to Markdown*\n\n")
            
            if not rtf_to_text_available:
                f.write("For full RTF conversion, please ensure striprtf is installed:\n")
                f.write("```\npip3 install striprtf\n```\n\n")
            
        print(f"✓ RTF to Markdown conversion complete: {output_file}")
        return output_file

    def _html_to_md(self, input_file, output_file):
        """Simple HTML to Markdown conversion for testing"""
        print(f"Converting HTML to Markdown: {input_file}")
        
        try:
            from bs4 import BeautifulSoup
            bs4_available = True
        except ImportError:
            bs4_available = False
            print("beautifulsoup4 not installed. Install with: pip3 install beautifulsoup4")
        
        if bs4_available:
            try:
                # Read HTML content
                with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
                    html_content = f.read()
                
                # Parse HTML
                soup = BeautifulSoup(html_content, 'html.parser')
                
                # Write to markdown
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    # Extract title from HTML
                    title_tag = soup.find('title')
                    if title_tag:
                        f.write(f'title: "{title_tag.get_text().strip()}"\n')
                    else:
                        f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Convert HTML to Markdown with proper formatting
                    body = soup.find('body') or soup
                    markdown_content = self._convert_html_to_markdown(body)
                    f.write(markdown_content)
                
                print(f"✓ HTML to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during HTML conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an HTML file converted to Markdown*\n\n")
            
            if not bs4_available:
                f.write("For full HTML conversion with image extraction, please ensure beautifulsoup4 is installed:\n")
                f.write("```\npip3 install beautifulsoup4\n```\n\n")
            
        print(f"✓ HTML to Markdown conversion complete: {output_file}")
        return output_file

    def _convert_html_to_markdown(self, element):
        """Convert HTML element to Markdown with proper formatting"""
        markdown = ""
        
        for child in element.children:
            if hasattr(child, 'name'):  # It's a tag
                tag_name = child.name.lower()
                
                if tag_name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    # Headers
                    level = int(tag_name[1])
                    markdown += f"{'#' * level} {child.get_text().strip()}\n\n"
                
                elif tag_name == 'p':
                    # Paragraphs
                    text = self._process_inline_elements(child)
                    if text.strip():
                        markdown += f"{text}\n\n"
                
                elif tag_name in ['ul', 'ol']:
                    # Lists
                    markdown += self._convert_list(child, ordered=(tag_name == 'ol'))
                    markdown += "\n"
                
                elif tag_name == 'blockquote':
                    # Blockquotes
                    text = self._process_inline_elements(child)
                    for line in text.split('\n'):
                        if line.strip():
                            markdown += f"> {line.strip()}\n"
                    markdown += "\n"
                
                elif tag_name in ['pre', 'code']:
                    # Code blocks and inline code
                    if tag_name == 'pre':
                        markdown += f"```\n{child.get_text()}\n```\n\n"
                    else:
                        markdown += f"`{child.get_text()}`"
                
                elif tag_name == 'a':
                    # Links
                    href = child.get('href', '')
                    text = child.get_text().strip()
                    if href and text:
                        markdown += f"[{text}]({href})"
                    else:
                        markdown += text
                
                elif tag_name in ['strong', 'b']:
                    # Bold text
                    text = self._process_inline_elements(child)
                    markdown += f"**{text}**"
                
                elif tag_name in ['em', 'i']:
                    # Italic text
                    text = self._process_inline_elements(child)
                    markdown += f"*{text}*"
                
                elif tag_name == 'img':
                    # Images
                    src = child.get('src', '')
                    alt = child.get('alt', 'Image')
                    title = child.get('title', '')
                    if src:
                        if title:
                            markdown += f'![{alt}]({src} "{title}")\n\n'
                        else:
                            markdown += f"![{alt}]({src})\n\n"
                
                elif tag_name in ['del', 's', 'strike']:
                    # Strikethrough text
                    text = self._process_inline_elements(child)
                    markdown += f"~~{text}~~"
                
                elif tag_name == 'sub':
                    # Subscript
                    text = child.get_text().strip()
                    markdown += f"~{text}~"
                
                elif tag_name == 'sup':
                    # Superscript
                    text = child.get_text().strip()
                    markdown += f"^{text}^"
                
                elif tag_name == 'mark':
                    # Highlight
                    text = self._process_inline_elements(child)
                    markdown += f"=={text}=="
                
                elif tag_name == 'table':
                    # Tables
                    markdown += self._convert_table(child)
                    markdown += "\n"
                
                elif tag_name in ['dl']:
                    # Definition lists
                    markdown += self._convert_definition_list(child)
                    markdown += "\n"
                
                elif tag_name == 'br':
                    # Line breaks
                    markdown += "\n"
                
                elif tag_name in ['div', 'span', 'section', 'article']:
                    # Container elements - process children
                    markdown += self._convert_html_to_markdown(child)
                
                elif tag_name == 'hr':
                    # Horizontal rule
                    markdown += "---\n\n"
                
                else:
                    # Other elements - just get text content
                    text = self._process_inline_elements(child)
                    if text.strip():
                        markdown += text
            
            else:  # It's text content
                text = str(child).strip()
                if text:
                    markdown += text
        
        return markdown

    def _process_inline_elements(self, element):
        """Process inline elements within a container"""
        result = ""
        
        for child in element.children:
            if hasattr(child, 'name'):  # It's a tag
                tag_name = child.name.lower()
                
                if tag_name == 'a':
                    href = child.get('href', '')
                    text = child.get_text().strip()
                    if href and text:
                        result += f"[{text}]({href})"
                    else:
                        result += text
                
                elif tag_name in ['strong', 'b']:
                    text = child.get_text().strip()
                    result += f"**{text}**"
                
                elif tag_name in ['em', 'i']:
                    text = child.get_text().strip()
                    result += f"*{text}*"
                
                elif tag_name == 'code':
                    result += f"`{child.get_text()}`"
                
                elif tag_name in ['del', 's', 'strike']:
                    text = child.get_text().strip()
                    result += f"~~{text}~~"
                
                elif tag_name == 'sub':
                    text = child.get_text().strip()
                    result += f"~{text}~"
                
                elif tag_name == 'sup':
                    text = child.get_text().strip()
                    result += f"^{text}^"
                
                elif tag_name == 'mark':
                    text = child.get_text().strip()
                    result += f"=={text}=="
                
                elif tag_name == 'br':
                    result += "\n"
                
                else:
                    result += child.get_text()
            else:
                result += str(child)
        
        return result

    def _convert_list(self, list_element, ordered=False):
        """Convert HTML list to Markdown"""
        markdown = ""
        items = list_element.find_all('li', recursive=False)
        
        for i, item in enumerate(items):
            if ordered:
                prefix = f"{i + 1}. "
            else:
                prefix = "- "
            
            text = self._process_inline_elements(item)
            # Handle nested lists
            nested_lists = item.find_all(['ul', 'ol'])
            for nested in nested_lists:
                nested_md = self._convert_list(nested, ordered=(nested.name == 'ol'))
                # Indent nested list
                nested_md = '\n'.join(f"  {line}" for line in nested_md.split('\n') if line.strip())
                text = text.replace(nested.get_text(), f"\n{nested_md}")
            
            markdown += f"{prefix}{text.strip()}\n"
        
        return markdown

    def _convert_table(self, table_element):
        """Convert HTML table to Markdown table"""
        rows = table_element.find_all('tr')
        if not rows:
            return ""
        
        markdown = "\n"
        header_processed = False
        
        for row in rows:
            cells = row.find_all(['th', 'td'])
            if not cells:
                continue
            
            # Process cells
            cell_texts = []
            for cell in cells:
                text = self._process_inline_elements(cell).strip()
                # Escape pipe characters in cell content
                text = text.replace('|', '\\|')
                cell_texts.append(text)
            
            # Create table row
            markdown += "| " + " | ".join(cell_texts) + " |\n"
            
            # Add header separator after first row (if it contains th elements or is first row)
            if not header_processed and (row.find('th') or rows.index(row) == 0):
                separator = "| " + " | ".join(['---'] * len(cells)) + " |\n"
                markdown += separator
                header_processed = True
        
        return markdown

    def _convert_definition_list(self, dl_element):
        """Convert HTML definition list to Markdown"""
        markdown = "\n"
        
        for child in dl_element.children:
            if hasattr(child, 'name'):
                if child.name == 'dt':
                    # Definition term
                    text = self._process_inline_elements(child).strip()
                    markdown += f"{text}\n"
                elif child.name == 'dd':
                    # Definition description
                    text = self._process_inline_elements(child).strip()
                    markdown += f": {text}\n\n"
        
        return markdown

    def _enhance_plain_text_to_markdown(self, content):
        """Convert plain text to structured Markdown with intelligent detection"""
        lines = content.split('\n')
        markdown_lines = []
        in_code_block = False
        
        for i, line in enumerate(lines):
            stripped = line.strip()
            
            # Skip empty lines (preserve them)
            if not stripped:
                markdown_lines.append('')
                continue
            
            # Detect and convert various patterns
            if not in_code_block:
                # Headers (lines ending with === or --- underneath)
                if i < len(lines) - 1:
                    next_line = lines[i + 1].strip()
                    if next_line and set(next_line) == {'='} and len(next_line) >= 3:
                        markdown_lines.append(f"# {stripped}")
                        lines[i + 1] = ''  # Skip the === line
                        continue
                    elif next_line and set(next_line) == {'-'} and len(next_line) >= 3:
                        markdown_lines.append(f"## {stripped}")
                        lines[i + 1] = ''  # Skip the --- line
                        continue
                
                # Numbered lists
                if re.match(r'^\d+\.\s+', stripped):
                    markdown_lines.append(stripped)  # Already valid Markdown
                    continue
                
                # Bullet points (various formats)
                if re.match(r'^[-*+•]\s+', stripped):
                    if not re.match(r'^[-*+]\s+', stripped):
                        # Convert various bullet characters to standard -
                        markdown_lines.append(re.sub(r'^[•]\s+', '- ', stripped))
                    else:
                        markdown_lines.append(stripped)  # Already valid
                    continue
                
                # Potential headers (ALL CAPS lines, or Title Case)
                if len(stripped) < 80 and (stripped.isupper() or stripped.istitle()) and not stripped.endswith('.'):
                    # Check if next line is content (not another title)
                    if i < len(lines) - 1 and lines[i + 1].strip() and not lines[i + 1].strip().isupper():
                        markdown_lines.append(f"## {stripped}")
                        continue
                
                # Code blocks (indented lines or lines with multiple code-like patterns)
                if (line.startswith('    ') or line.startswith('\t') or 
                    (re.search(r'[{}()[\]<>]', stripped) and '=' in stripped and (';' in stripped or '{' in stripped))):
                    if not in_code_block:
                        markdown_lines.append('```')
                        in_code_block = True
                    markdown_lines.append(line.rstrip())
                    continue
                else:
                    if in_code_block:
                        markdown_lines.append('```')
                        in_code_block = False
                
                # URLs - make them clickable
                if re.match(r'^https?://', stripped):
                    markdown_lines.append(f"<{stripped}>")
                    continue
                
                # Email addresses
                if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', stripped):
                    markdown_lines.append(f"<{stripped}>")
                    continue
                
                # Horizontal rules (series of dashes, asterisks, or underscores)
                if re.match(r'^[-*_]{3,}$', stripped):
                    markdown_lines.append('---')
                    continue
            
            # Default: treat as paragraph text
            markdown_lines.append(stripped)
        
        # Close code block if still open
        if in_code_block:
            markdown_lines.append('```')
        
        # Join lines and clean up multiple empty lines
        result = '\n'.join(markdown_lines)
        result = re.sub(r'\n{3,}', '\n\n', result)  # Max 2 consecutive newlines
        
        return result + '\n\n'

    def _xlsx_to_md(self, input_file, output_file):
        """Simple XLSX to Markdown conversion for testing"""
        print(f"Converting XLSX to Markdown: {input_file}")
        
        try:
            from openpyxl import load_workbook
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
            print("openpyxl not installed. Install with: pip3 install openpyxl")
        
        if openpyxl_available:
            try:
                # Load the workbook
                workbook = load_workbook(input_file, data_only=True)
                
                with open(output_file, 'w', encoding='utf-8') as f:
                    # YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "MarkdownMagic"\n')
                    f.write("---\n\n")
                    
                    # Don't add redundant title - let sheet names provide document structure
                    
                    # Process each worksheet
                    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                        sheet = workbook[sheet_name]
                        
                        # Add sheet as main heading
                        f.write(f"# {sheet_name}\n\n")
                        
                        # Get the maximum row and column
                        max_row = sheet.max_row
                        max_col = sheet.max_column
                        
                        if max_row > 0 and max_col > 0:
                            # Create table header
                            header_row = []
                            for col in range(1, max_col + 1):
                                cell_value = sheet.cell(row=1, column=col).value
                                header_row.append(str(cell_value) if cell_value is not None else '')
                            
                            f.write("| " + " | ".join(header_row) + " |\n")
                            f.write("| " + " | ".join(["---"] * max_col) + " |\n")
                            
                            # Create table rows
                            for row in range(2, max_row + 1):
                                row_data = []
                                for col in range(1, max_col + 1):
                                    cell_value = sheet.cell(row=row, column=col).value
                                    row_data.append(str(cell_value) if cell_value is not None else '')
                                
                                f.write("| " + " | ".join(row_data) + " |\n")
                            
                            f.write("\n")
                        else:
                            f.write("*This sheet is empty*\n\n")
                        
                        # Add separator between sheets
                        if sheet_index < len(workbook.sheetnames) - 1:
                            f.write("---\n\n")
                
                print(f"✓ XLSX to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during XLSX conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "MarkdownMagic"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an Excel file converted to Markdown*\n\n")
            
            if not openpyxl_available:
                f.write("For full Excel conversion with chart extraction, please ensure openpyxl is installed:\n")
                f.write("```\npip3 install openpyxl\n```\n\n")
            
        print(f"✓ XLSX to Markdown conversion complete: {output_file}")
        return output_file

    def _image_to_md(self, input_file, output_file):
        """Convert image files to Markdown with OCR"""
        print(f"Converting image to Markdown: {input_file}")
        
        if not self.image_processor:
            # Fallback: create basic markdown without OCR
            with open(output_file, 'w', encoding='utf-8') as f:
                # Add YAML frontmatter
                f.write("---\n")
                f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                f.write(f'source: "{input_file}"\n')
                f.write('converter: "MarkdownMagic"\n')
                f.write('type: "image"\n')
                f.write("---\n\n")
                
                # Add title
                f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                
                # Add image reference
                f.write(f"![{os.path.basename(input_file)}]({input_file})\n\n")
                
                f.write("*Image processing and OCR require image_processor.py*\n\n")
            
            print(f"✓ Basic image to Markdown conversion complete: {output_file}")
            return output_file
        
        try:
            # Use image processor for full conversion with OCR
            result_file = self.image_processor.process_image_file(input_file, output_file)
            return result_file
            
        except Exception as e:
            print(f"! Error during image conversion: {e}")
            # Fall back to basic conversion
            with open(output_file, 'w', encoding='utf-8') as f:
                # Add YAML frontmatter
                f.write("---\n")
                f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                f.write(f'source: "{input_file}"\n')
                f.write('converter: "MarkdownMagic"\n')
                f.write('type: "image"\n')
                f.write("---\n\n")
                
                # Add title
                f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                
                # Add image reference
                f.write(f"![{os.path.basename(input_file)}]({input_file})\n\n")
                
                f.write(f"*Error during image processing: {str(e)}*\n\n")
            
            print(f"✓ Fallback image to Markdown conversion complete: {output_file}")
            return output_file
    
    def _find_common_lines(self, line_groups):
        """Find lines that appear in multiple page groups (headers/footers)"""
        if not line_groups or len(line_groups) < 2:
            return []
        
        # Count occurrences of each line
        line_counts = {}
        total_groups = len(line_groups)
        
        for group in line_groups:
            for line in group:
                line = line.strip()
                if line and len(line) > 3:  # Ignore very short lines
                    line_counts[line] = line_counts.get(line, 0) + 1
        
        # Return lines that appear in at least 50% of pages
        threshold = max(2, total_groups // 2)
        common_lines = [line for line, count in line_counts.items() if count >= threshold]
        return common_lines
    
    def _remove_headers_footers(self, text, common_headers, common_footers):
        """Remove common headers and footers from page text"""
        if not text.strip():
            return text
            
        lines = text.split('\n')
        
        # Remove header lines
        for header_line in common_headers:
            # Remove from beginning of text
            while lines and lines[0].strip() == header_line.strip():
                lines.pop(0)
        
        # Remove footer lines  
        for footer_line in common_footers:
            # Remove from end of text
            while lines and lines[-1].strip() == footer_line.strip():
                lines.pop()
        
        # Also remove lines that contain common copyright/title patterns
        filtered_lines = []
        for line in lines:
            line_clean = line.strip().lower()
            # Skip lines that look like headers/footers
            if (line_clean.startswith('©') or 
                'copyright' in line_clean or 
                'all rights reserved' in line_clean or
                line_clean.startswith('page ') or
                line_clean.endswith(' page') or
                len(line_clean) < 3):
                continue
            filtered_lines.append(line)
        
        return '\n'.join(filtered_lines)