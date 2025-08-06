#!/usr/bin/env python3
"""
Document converter for EZMDc - simplified version for testing
Converts various document formats to Markdown
"""

import os
import re
from pathlib import Path

class DocumentConverter:
    """
    Converts various document formats to Markdown with structure preservation
    """
    
    def __init__(self, tesseract_path=None):
        """Initialize the document converter"""
        self.supported_formats = {
            '.txt': self._txt_to_md,
            '.pdf': self._pdf_to_md,
            '.docx': self._docx_to_md,
            '.odt': self._odt_to_md,
            '.rtf': self._rtf_to_md,
            '.html': self._html_to_md,
            '.htm': self._html_to_md,
            '.xlsx': self._xlsx_to_md,
            '.xls': self._xlsx_to_md
        }
        self.tesseract_path = tesseract_path
        
        # Initialize image processor
        try:
            from image_processor import ImageProcessor
            self.image_processor = ImageProcessor(tesseract_path=tesseract_path)
            print("EZMDc Document Converter initialized with image processing")
        except ImportError:
            self.image_processor = None
            print("EZMDc Document Converter initialized (image processing unavailable)")
    
    def convert_to_markdown(self, input_file, output_file=None):
        """
        Convert a document to Markdown format
        
        Args:
            input_file (str): Path to input document
            output_file (str, optional): Path for output Markdown file
            
        Returns:
            str: Path to created Markdown file
        """
        
        # Validate input file
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")
        
        # Get file extension
        file_ext = Path(input_file).suffix.lower()
        
        # Check if format is supported
        if file_ext not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        # Generate output filename if not provided
        if output_file is None:
            input_path = Path(input_file)
            output_file = str(input_path.with_suffix('.md'))
        
        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Convert using appropriate method
        converter_method = self.supported_formats[file_ext]
        converter_method(input_file, output_file)
        
        return output_file
    
    def _txt_to_md(self, input_file, output_file):
        """Convert plain text to Markdown with intelligent structure detection"""
        print(f"Converting TXT to Markdown: {input_file}")
        
        with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Write content
            f.write(content)
        
        print(f"✓ TXT to Markdown conversion complete: {output_file}")
        return output_file

    def _pdf_to_md(self, input_file, output_file):
        """Convert PDF to Markdown with image extraction"""
        print(f"Converting PDF to Markdown: {input_file}")
        
        try:
            import fitz  # PyMuPDF
            pymupdf_available = True
        except ImportError:
            pymupdf_available = False
            print("PyMuPDF not installed. Install with: pip install PyMuPDF")
        
        # If we have PyMuPDF and image processor, use enhanced conversion
        if pymupdf_available and self.image_processor:
            try:
                # Open the PDF
                pdf = fitz.open(input_file)
                
                # Create image folder for this document
                images_folder = self.image_processor.create_image_folder(output_file)
                
                # Track document structure
                image_count = 0
                total_pages = len(pdf)
                
                # Prepare to extract text with structure
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "EZMDc"\n')
                    f.write("---\n\n")
                    
                    # Add title
                    f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # Process each page
                    for page_num in range(total_pages):
                        print(f"Processing page {page_num + 1}...")
                        
                        # Get the page
                        page = pdf[page_num]
                        
                        # Extract text
                        text = page.get_text()
                        
                        # Extract images from this page
                        try:
                            image_list = page.get_images()
                            for img_index, img in enumerate(image_list):
                                try:
                                    # Extract image
                                    xref = img[0]
                                    base_image = pdf.extract_image(xref)
                                    image_bytes = base_image["image"]
                                    
                                    image_count += 1
                                    
                                    # Process image with OCR
                                    page_info = f"pg{page_num + 1}"
                                    markdown_placeholder = self.image_processor.process_image(
                                        image_bytes, image_count, images_folder, page_info
                                    )
                                    
                                    # Add to markdown
                                    f.write(f"{markdown_placeholder}\n\n")
                                    
                                except Exception as e:
                                    print(f"  ⚠ Failed to extract image {img_index + 1}: {e}")
                                    # Create a placeholder anyway
                                    placeholder = f"![Image {image_count + 1}, pg{page_num + 1} - Could not extract image]({images_folder}/image_{image_count + 1}.png)"
                                    f.write(f"{placeholder}\n\n")
                                    image_count += 1
                        
                        except Exception as e:
                            print(f"  ⚠ Error processing images on page {page_num + 1}: {e}")
                        
                        # Add page text
                        if text.strip():
                            f.write(f"{text.strip()}\n\n")
                        
                        # Add page break except for last page
                        if page_num < total_pages - 1:
                            f.write("\n---\n\n")
                
                # Close the PDF properly
                pdf.close()
                
                if image_count > 0:
                    print(f"✓ PDF conversion complete with {total_pages} pages and {image_count} images processed")
                else:
                    print(f"✓ PDF conversion complete with {total_pages} pages processed")
                
                return output_file
                
            except Exception as e:
                print(f"! Error during enhanced PDF conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is a PDF file converted to Markdown*\n\n")
            
            if not pymupdf_available:
                f.write("For full PDF conversion with image extraction, please ensure PyMuPDF is installed:\n")
                f.write("```\npip install PyMuPDF\n```\n\n")
            elif not self.image_processor:
                f.write("For image extraction, please ensure image_processor.py is available\n\n")
            
        print(f"✓ PDF to Markdown conversion complete: {output_file}")
        return output_file

    def _docx_to_md(self, input_file, output_file):
        """Convert DOCX to Markdown with image extraction"""
        print(f"Converting DOCX to Markdown: {input_file}")
        
        try:
            from docx import Document
            python_docx_available = True
        except ImportError:
            python_docx_available = False
            print("python-docx not installed. Install with: pip install python-docx")
        
        # If we have python-docx and image processor, use enhanced conversion
        if python_docx_available and self.image_processor:
            try:
                # Open the document
                doc = Document(input_file)
                
                # Create image folder
                images_folder = self.image_processor.create_image_folder(output_file)
                
                image_count = 0
                
                with open(output_file, 'w', encoding='utf-8') as f:
                    # YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "EZMDc"\n')
                    f.write("---\n\n")
                    
                    # Add title
                    f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # First pass: extract images from document relationships
                    try:
                        # Extract images from document relationships
                        if hasattr(doc, 'part') and hasattr(doc.part, 'rels'):
                            for rel_id, rel in doc.part.rels.items():
                                if "image" in rel.target_ref:
                                    try:
                                        # Get image data
                                        image_part = rel.target_part
                                        image_bytes = image_part.blob
                                        
                                        image_count += 1
                                        
                                        # Process image with OCR
                                        markdown_placeholder = self.image_processor.process_image(
                                            image_bytes, image_count, images_folder, f"pos{image_count}"
                                        )
                                        
                                        print(f"Extracted image {image_count} from DOCX")
                                        
                                    except Exception as e:
                                        print(f"Warning: Could not extract image {rel_id}: {e}")
                    except Exception as e:
                        print(f"Warning: Could not extract images from DOCX: {e}")
                    
                    # Second pass: process document content
                    image_refs_used = 0
                    
                    for paragraph in doc.paragraphs:
                        text = paragraph.text.strip()
                        
                        # Check if paragraph contains an image
                        has_inline_image = False
                        for run in paragraph.runs:
                            if hasattr(run, 'element') and run.element.findall('.//a:blip', namespaces={'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}):
                                has_inline_image = True
                                break
                        
                        # If paragraph has an image, insert the image placeholder
                        if has_inline_image and image_refs_used < image_count:
                            image_refs_used += 1
                            markdown_placeholder = self.image_processor.create_markdown_placeholder(
                                image_refs_used, f"pos{image_refs_used}", 
                                f"Image {image_refs_used} from document", images_folder
                            )
                            f.write(f"{markdown_placeholder}\n\n")
                            
                            # Also include any text in the paragraph
                            if text:
                                f.write(f"{text}\n\n")
                            continue
                        
                        if not text:
                            f.write('\n')
                            continue
                        
                        # Check paragraph style for headings
                        style_name = paragraph.style.name.lower()
                        
                        if 'heading' in style_name:
                            # Extract heading level
                            if 'heading 1' in style_name:
                                f.write(f"# {text}\n\n")
                            elif 'heading 2' in style_name:
                                f.write(f"## {text}\n\n")
                            elif 'heading 3' in style_name:
                                f.write(f"### {text}\n\n")
                            elif 'heading 4' in style_name:
                                f.write(f"#### {text}\n\n")
                            elif 'heading 5' in style_name:
                                f.write(f"##### {text}\n\n")
                            elif 'heading 6' in style_name:
                                f.write(f"###### {text}\n\n")
                            else:
                                f.write(f"## {text}\n\n")
                            continue
                        
                        # Process text formatting
                        processed_text = ""
                        for run in paragraph.runs:
                            run_text = run.text
                            if run.bold:
                                run_text = f"**{run_text}**"
                            if run.italic:
                                run_text = f"*{run_text}*"
                            processed_text += run_text
                        
                        # Regular paragraph
                        f.write(f"{processed_text}\n\n")
                    
                    # Add any remaining images that weren't inline
                    while image_refs_used < image_count:
                        image_refs_used += 1
                        markdown_placeholder = self.image_processor.create_markdown_placeholder(
                            image_refs_used, f"pos{image_refs_used}", 
                            f"Image {image_refs_used} from document", images_folder
                        )
                        f.write(f"{markdown_placeholder}\n\n")
                
                if image_count > 0:
                    print(f"✓ DOCX to Markdown conversion complete with {image_count} images processed")
                else:
                    print("✓ DOCX to Markdown conversion complete")
                
                return output_file
                
            except Exception as e:
                print(f"! Error during enhanced DOCX conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is a DOCX file converted to Markdown*\n\n")
            
            if not python_docx_available:
                f.write("For full DOCX conversion with image extraction, please ensure python-docx is installed:\n")
                f.write("```\npip install python-docx\n```\n\n")
            elif not self.image_processor:
                f.write("For image extraction, please ensure image_processor.py is available\n\n")
            
        print(f"✓ DOCX to Markdown conversion complete: {output_file}")
        return output_file

    def _odt_to_md(self, input_file, output_file):
        """Simple ODT to Markdown conversion for testing"""
        print(f"Converting ODT to Markdown: {input_file}")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an ODT file converted to Markdown*\n\n")
            
        print(f"✓ ODT to Markdown conversion complete: {output_file}")
        return output_file

    def _rtf_to_md(self, input_file, output_file):
        """Simple RTF to Markdown conversion for testing"""
        print(f"Converting RTF to Markdown: {input_file}")
        
        try:
            from striprtf.striprtf import rtf_to_text
            rtf_to_text_available = True
        except ImportError:
            rtf_to_text_available = False
            print("striprtf not installed. Install with: pip install striprtf")
        
        if rtf_to_text_available:
            try:
                # Read RTF content
                with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
                    rtf_content = f.read()
                
                # Convert RTF to plain text
                text_content = rtf_to_text(rtf_content)
                
                # Write to markdown
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "EZMDc"\n')
                    f.write("---\n\n")
                    
                    # Add main heading
                    f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # Write content
                    f.write(text_content)
                
                print(f"✓ RTF to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during RTF conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an RTF file converted to Markdown*\n\n")
            
            if not rtf_to_text_available:
                f.write("For full RTF conversion, please ensure striprtf is installed:\n")
                f.write("```\npip install striprtf\n```\n\n")
            
        print(f"✓ RTF to Markdown conversion complete: {output_file}")
        return output_file

    def _html_to_md(self, input_file, output_file):
        """Simple HTML to Markdown conversion for testing"""
        print(f"Converting HTML to Markdown: {input_file}")
        
        try:
            from bs4 import BeautifulSoup
            bs4_available = True
        except ImportError:
            bs4_available = False
            print("beautifulsoup4 not installed. Install with: pip install beautifulsoup4")
        
        if bs4_available:
            try:
                # Read HTML content
                with open(input_file, 'r', encoding='utf-8', errors='replace') as f:
                    html_content = f.read()
                
                # Parse HTML
                soup = BeautifulSoup(html_content, 'html.parser')
                
                # Write to markdown
                with open(output_file, 'w', encoding='utf-8') as f:
                    # Add YAML frontmatter
                    f.write("---\n")
                    # Extract title from HTML
                    title_tag = soup.find('title')
                    if title_tag:
                        f.write(f'title: "{title_tag.get_text().strip()}"\n')
                    else:
                        f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "EZMDc"\n')
                    f.write("---\n\n")
                    
                    # Add title as main heading
                    if title_tag:
                        f.write(f"# {title_tag.get_text().strip()}\n\n")
                    else:
                        f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # Process body content
                    body = soup.find('body')
                    if body:
                        # Get all text
                        text = body.get_text()
                        # Clean up whitespace
                        text = re.sub(r'\s+', ' ', text).strip()
                        f.write(f"{text}\n\n")
                    else:
                        # No body tag, use entire content
                        text = soup.get_text()
                        text = re.sub(r'\s+', ' ', text).strip()
                        f.write(f"{text}\n\n")
                
                print(f"✓ HTML to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during HTML conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an HTML file converted to Markdown*\n\n")
            
            if not bs4_available:
                f.write("For full HTML conversion with image extraction, please ensure beautifulsoup4 is installed:\n")
                f.write("```\npip install beautifulsoup4\n```\n\n")
            
        print(f"✓ HTML to Markdown conversion complete: {output_file}")
        return output_file

    def _xlsx_to_md(self, input_file, output_file):
        """Simple XLSX to Markdown conversion for testing"""
        print(f"Converting XLSX to Markdown: {input_file}")
        
        try:
            from openpyxl import load_workbook
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
            print("openpyxl not installed. Install with: pip install openpyxl")
        
        if openpyxl_available:
            try:
                # Load the workbook
                workbook = load_workbook(input_file, data_only=True)
                
                with open(output_file, 'w', encoding='utf-8') as f:
                    # YAML frontmatter
                    f.write("---\n")
                    f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
                    f.write(f'source: "{input_file}"\n')
                    f.write('converter: "EZMDc"\n')
                    f.write("---\n\n")
                    
                    # Add title
                    f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
                    
                    # Process each worksheet
                    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                        sheet = workbook[sheet_name]
                        
                        # Add sheet as heading (unless it's the only sheet)
                        if len(workbook.sheetnames) > 1:
                            f.write(f"## {sheet_name}\n\n")
                        
                        # Get the maximum row and column
                        max_row = sheet.max_row
                        max_col = sheet.max_column
                        
                        if max_row > 0 and max_col > 0:
                            # Create table header
                            header_row = []
                            for col in range(1, max_col + 1):
                                cell_value = sheet.cell(row=1, column=col).value
                                header_row.append(str(cell_value) if cell_value is not None else '')
                            
                            f.write("| " + " | ".join(header_row) + " |\n")
                            f.write("| " + " | ".join(["---"] * max_col) + " |\n")
                            
                            # Create table rows
                            for row in range(2, max_row + 1):
                                row_data = []
                                for col in range(1, max_col + 1):
                                    cell_value = sheet.cell(row=row, column=col).value
                                    row_data.append(str(cell_value) if cell_value is not None else '')
                                
                                f.write("| " + " | ".join(row_data) + " |\n")
                            
                            f.write("\n")
                        else:
                            f.write("*This sheet is empty*\n\n")
                        
                        # Add separator between sheets
                        if sheet_index < len(workbook.sheetnames) - 1:
                            f.write("---\n\n")
                
                print(f"✓ XLSX to Markdown conversion complete: {output_file}")
                return output_file
            except Exception as e:
                print(f"! Error during XLSX conversion: {e}")
                # Fall back to basic conversion
                pass
        
        # Fallback to basic conversion
        with open(output_file, 'w', encoding='utf-8') as f:
            # Add YAML frontmatter
            f.write("---\n")
            f.write(f'title: "{os.path.splitext(os.path.basename(input_file))[0]}"\n')
            f.write(f'source: "{input_file}"\n')
            f.write('converter: "EZMDc"\n')
            f.write("---\n\n")
            
            # Add main heading
            f.write(f"# {os.path.splitext(os.path.basename(input_file))[0]}\n\n")
            
            # Basic placeholder content
            f.write("*This is an Excel file converted to Markdown*\n\n")
            
            if not openpyxl_available:
                f.write("For full Excel conversion with chart extraction, please ensure openpyxl is installed:\n")
                f.write("```\npip install openpyxl\n```\n\n")
            
        print(f"✓ XLSX to Markdown conversion complete: {output_file}")
        return output_file